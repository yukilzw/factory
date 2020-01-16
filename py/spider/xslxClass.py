import os
import xlsxwriter
import numpy as np
import pandas as pd
from openpyxl import load_workbook


class MxlsxWB():
    def __init__(self, workpath=os.getcwd(), filename=None):
        self.workpath = workpath  # default current path
        self.filename = filename

    # set workpath
    def set_path(self, workpath): 
        self.workpath = workpath
        os.chdir(self.workpath)

    # get file base information
    def get_fileinfo(self):
        # print(self.filename)
        print("=" * 30, "FILE INFO", "=" * 30)  # cut line
        self.wb = load_workbook(filename=self.filename)
        self.sheetnames = self.wb.get_sheet_names()

        print("文件" + self.filename + "共包含", len(self.sheetnames), "个工作表")
        print("表名为：", end=" ")
        for name in self.sheetnames:
            print(name, end=" ")
        print("\n")
        print("=" * 30, "END FILE INFO", "=" * 30)  # cut line

    # choose work table
    def choose_sheet(self, sheetname=None):
        if sheetname == None:
            self.sheetname = self.sheetnames[0]

        self.sheetname = sheetname
        self.worksheet = self.wb[self.sheetname]

    # get work table information
    def get_sheetinfo(self):

        print("=" * 30, self.sheetname, "=" * 30)  # cut line

        self.num_of_rows = len(list(self.worksheet.rows))
        self.num_of_cols = len(list(self.worksheet.columns))

        print("行数：", self.num_of_rows)
        print("列数：", self.num_of_cols)
        print("列名：", MxlsxWB.get_rowdata(self, rownum=1))

        print("=" * 30, self.sheetname, "=" * 30)  # cut line

    '''
    openpyxl —— data search
    '''

    # get each row data
    def get_rowdata(self, rownum):
        rowdata = []
        for row in self.worksheet.iter_rows(min_row=rownum, max_row=rownum, max_col=self.num_of_cols):
            for cell in row:
                rowdata.append(cell.value)
        # print(rowdata)
        return rowdata

    # get each column data
    def get_coldata(self, colnum):
        coldata = []
        for col in self.worksheet.iter_cols(min_row=colnum, max_row=colnum, max_col=self.num_of_rows):
            for cell in col:
                coldata.append(cell.value)
        # print(coldata)
        return col

    def get_areadata(self, min_row, max_row, min_col, max_col):
        print("=" * 30, "区域数据", "=" * 30)

        # Create empty Matrix whith type to be string
        areadata = np.matrix(np.zeros((max_row - min_row + 1, max_col - min_col + 1)), dtype=str)
        for col in self.worksheet.iter_cols(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in col:
                col_index = cell.col_idx
                row_index = cell.row
                areadata[row_index - min_row, col_index - min_col] = cell.value
        print(areadata)

        print("=" * 30, "区域数据", "=" * 30)  # 分割线

        return areadata

    '''基于xlsxwriter——数据的写入'''

    def create_workbook(self, wb_name):
        if not '.xlsx' in wb_name:
            self.wb = xlsxwriter.Workbook(wb_name + '.xlsx')
        self.wb = xlsxwriter.Workbook(wb_name)

    def create_worksheet(self, ws_name):
        self.worksheet = self.wb.add_worksheet(ws_name)

    def add_col_names(self, col_names):
        self.num_of_cols = len(col_names)
        for i in range(self.num_of_cols):
            self.worksheet.write(0, i, col_names[i])

    def add_coldata(self, data, colx):
        self.num_of_rows = len(data)
        for row in range(len(data)):
            self.worksheet.write(row + 1, colx - 1, data[row])

    def add_rowdata(self, data, rowx):
        for col in range(self.num_of_cols):
            self.worksheet.write(rowx - 1, col, data[col])

    def save(self):
        self.wb.close()

    '''pandas'''

    def read_by_pandas(self, filename=None):
        if filename == None:
            filename = self.filename
        df = pd.read_excel(filename)

        print("=" * 10, "DataFrame From " + filename + ":", "=" * 10)
        print(df)
        print("=" * 10, "DataFrame From " + filename + ":", "=" * 10)

        return df

    def write_by_pandas(self, df, new_filename, new_sheetname):
        df.to_excel(new_filename, sheetname=new_sheetname)


if __name__ == '__main__':
    Demo = MxlsxWB(filename='pandas_simple.xlsx')
    Demo.set_path("Myxlsxdata")

    Demo.get_fileinfo()
    Demo.choose_sheet('豆瓣图书')
    Demo.get_sheetinfo()
    Demo.get_areadata(2, 3, 2, 3)

    Demo.create_workbook('Mxlsxclass.xlsx')
    Demo.create_worksheet('s1')
    Demo.add_col_names(['col1', 'col2'])
    Demo.add_coldata([1, 2, 3, 4, 5], 1)
    Demo.add_coldata([2, 3, 4, 5, 6], 2)
    Demo.save()

    Demo.read_by_pandas('Mxlsxclass.xlsx')